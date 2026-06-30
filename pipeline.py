from grading import generate_answer_key, grading_wrapper, build_bubble_index, get_bubble_coordinates, print_report, debug_view
import fitz
from pathlib import Path
import cv2
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

input_dir = Path("./input")
output_dir = Path("./output")
template_path = Path("./templates/CMS_mc_template.json")
answer_key_path = Path("./samples/answers/1.png")

template = json.load(open(template_path, "r", encoding="utf-8"))
index = build_bubble_index(template)
answer_key_img = cv2.imread(answer_key_path)
answer_key = generate_answer_key(answer_key_img, get_bubble_coordinates(answer_key_img, template), index)

input_dir.mkdir(parents=True, exist_ok=True)
output_dir.mkdir(parents=True, exist_ok=True)

for idx, file in enumerate(input_dir.glob("*.pdf")):
  doc = fitz.open(file)

  for page_num in range(len(doc)):
    page = doc.load_page(page_num)
    pix = page.get_pixmap()
    
    # 4. Save the pixmap as an image file (supports PNG, JPG, etc.)
    output_path = f"{output_dir}/{idx}_page_{page_num + 1}.png"
    pix.save(output_path)
    print(f"Saved: {output_path}")

results = []
for image in output_dir.glob("*.png"):
  img = cv2.imread(image)
  result = grading_wrapper(img, answer_key, template, index, fill_threshold=0.75)
  # print_report(result)
  # debug_view(img, get_bubble_coordinates(img,template), index)
  # print(result["metadata"]["Class"])
  results.append(result)

results_sorted = sorted(
    results, 
    key=lambda x: (x["metadata"]["Class"], x["metadata"]["Class Number"])
)

def generate_excel_dashboard(all_classes_data, output_filename="Exam_Report.xlsx"):
    """
    Generates a beautifully formatted multi-sheet Excel workbook from multi-class raw data.
    Includes a Master Dashboard with charts, individual class rosters with rankings, 
    and a cohort-wide Item Analysis diagnostic sheet.
    """
    if not all_classes_data:
        print("No data provided.")
        return

    # ----------------------------------------------------
    # 1. INITIAL DATA PROCESSING & AGGREGATION
    # ----------------------------------------------------
    # Group students by their class
    classes_dict = {}
    active_qs = set()
    expected_answers = {}

    for student in all_classes_data:
        cls = student['metadata'].get('Class', 'Unknown')
        if cls not in classes_dict:
            classes_dict[cls] = []
        classes_dict[cls].append(student)
        
        # Track evaluated questions and save answer key
        for q, res in student.get('results', {}).items():
            active_qs.add(q)
            if q not in expected_answers and 'expected' in res:
                expected_answers[q] = res['expected']

    active_qs = sorted(list(active_qs))
    sorted_classes = sorted(classes_dict.keys())

    # Keep track of row bounds for formulas
    class_sheet_ranges = {}

    # ----------------------------------------------------
    # 2. CREATE WORKBOOK AND STYLES
    # ----------------------------------------------------
    wb = openpyxl.Workbook()
    active_sheet = wb.active
    if active_sheet is not None:
        wb.remove(active_sheet)

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=10, italic=True)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    sub_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    summary_fill = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
    warn_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid") # Soft yellow

    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7')
    )
    double_bottom_border = Border(top=Side(style='thin', color='BDC3C7'), bottom=Side(style='double', color='2C3E50'))

    # ----------------------------------------------------
    # 3. TAB 1: MASTER DASHBOARD
    # ----------------------------------------------------
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash["A1"] = "COHORT PERFORMANCE DASHBOARD"
    ws_dash["A1"].font = title_font
    
    ws_dash["A3"] = "Cohort Global Metrics"
    ws_dash["A3"].font = section_font
    
    # We will build formulas dynamically pointing to class sheets later
    dash_metrics = [
        ["Metric", "Value"],
        ["Total Classes Tracked", len(sorted_classes)],
        ["Total Students Assessed", ""], 
        ["Cohort Average Score", ""],
        ["Cohort Median Score", ""],
        ["Highest Individual Score", ""],
        ["Lowest Individual Score", ""],
        ["Overall Pass Rate (>=50%)", ""]
    ]
    
    for r_idx, row in enumerate(dash_metrics, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 4:
                cell.font = header_font; cell.fill = header_fill
            else:
                cell.font = regular_font; cell.border = thin_border
                if c_idx == 2: cell.alignment = Alignment(horizontal="right")

    # Class Breakdown Table Headers
    ws_dash["A13"] = "Class Comparison Breakdown"
    ws_dash["A13"].font = section_font
    comp_headers = ["Class", "Registered Students", "Class Average", "Highest Score", "Pass Rate"]
    for c_idx, h in enumerate(comp_headers, start=1):
        cell = ws_dash.cell(row=14, column=c_idx, value=h)
        cell.font = header_font; cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

    # ----------------------------------------------------
    # 4. TABS 2+: INDIVIDUAL CLASS ROSTERS
    # ----------------------------------------------------
    for cls_name in sorted_classes:
        students = classes_dict[cls_name]
        ws = wb.create_sheet(title=f"Class {cls_name}")
        ws.views.sheetView[0].showGridLines = True
        
        ws["A1"] = f"Student Performance Roster - Class {cls_name}"
        ws["A1"].font = title_font
        
        headers = ["Class", "Class Number", "Reg. Number", "Subject Code", "Score", "Total", "Percentage", "Class Rank"]
        for q in active_qs:
            headers.append(f"Q{q}")
            
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Reference Key Row
        key_row = ["KEY", "-", "-", "-", "-", "-", "-", "-"] + [expected_answers.get(q, '-') for q in active_qs]
        for c_idx, val in enumerate(key_row, start=1):
            cell = ws.cell(row=5, column=c_idx, value=val)
            cell.font = italic_font; cell.fill = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
            cell.alignment = Alignment(horizontal="center"); cell.border = thin_border

        start_row = 6
        end_row = start_row + len(students) - 1
        class_sheet_ranges[cls_name] = (start_row, end_row)

        for s_idx, s in enumerate(students):
            r = start_row + s_idx
            is_zebra = (s_idx % 2 == 1)
            current_fill = zebra_fill if is_zebra else PatternFill(fill_type=None)
            meta = s['metadata']
            
            ws.cell(row=r, column=1, value=str(meta.get('Class'))).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=2, value=str(meta.get('Class Number'))).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=3, value=str(meta.get('Registration Number'))).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=4, value=str(meta.get('Subject Code'))).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=5, value=s['score']).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=6, value=s['total']).alignment = Alignment(horizontal="right")
            
            pct_cell = ws.cell(row=r, column=7, value=f"=E{r}/F{r}")
            pct_cell.number_format = '0.0%'; pct_cell.alignment = Alignment(horizontal="right")
            
            # Auto Rank Formula inside this class
            rank_cell = ws.cell(row=r, column=8, value=f"=RANK(E{r}, E${start_row}:E${end_row})")
            rank_cell.alignment = Alignment(horizontal="center")
            
            # Map item answers
            for q_idx, q in enumerate(active_qs, start=9):
                ans_val = s['answers'].get(q)
                q_cell = ws.cell(row=r, column=q_idx)
                q_cell.alignment = Alignment(horizontal="center")
                
                if ans_val is None:
                    q_cell.value = "BLANK"; q_cell.fill = warn_fill; q_cell.font = italic_font
                elif ans_val == "AMBIGUOUS":
                    q_cell.value = "AMBIG"; q_cell.fill = warn_fill; q_cell.font = bold_font
                else:
                    q_cell.value = ans_val
                    if ans_val != expected_answers.get(q):
                        q_cell.font = Font(name=font_family, size=11, color="922B21", bold=True) # Soft crimson red text for errors
                    else:
                        q_cell.font = regular_font

            for c_idx in range(1, len(headers) + 1):
                c = ws.cell(row=r, column=c_idx); c.border = thin_border
                if c.fill.fill_type is None: c.fill = current_fill

        # Summary Metrics Row per Class
        avg_row = end_row + 1
        ws.cell(row=avg_row, column=1, value="Class Average").font = bold_font
        ws.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=4)
        
        avg_score_cell = ws.cell(row=avg_row, column=5, value=f"=AVERAGE(E{start_row}:E{end_row})")
        avg_score_cell.font = bold_font; avg_score_cell.number_format = '0.0'; avg_score_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=avg_row, column=6, value="")
        
        avg_pct_cell = ws.cell(row=avg_row, column=7, value=f"=AVERAGE(G{start_row}:G{end_row})")
        avg_pct_cell.font = bold_font; avg_pct_cell.number_format = '0.0%'; avg_pct_cell.alignment = Alignment(horizontal="right")
        
        pass_rate_cell = ws.cell(row=avg_row, column=8, value=f'=COUNTIF(G{start_row}:G{end_row},">=50%")/COUNT(G{start_row}:G{end_row})')
        pass_rate_cell.font = bold_font; pass_rate_cell.number_format = '0.0%'; pass_rate_cell.alignment = Alignment(horizontal="center")
        
        for c_idx in range(1, len(headers) + 1):
            c = ws.cell(row=avg_row, column=c_idx); c.border = double_bottom_border; c.fill = summary_fill

    # ----------------------------------------------------
    # 5. RETROFIT DASHBOARD FORMULAS & BREAKDOWN
    # ----------------------------------------------------
    total_stud_segments = ", ".join([f"'Class {c}'!B$6:B${class_sheet_ranges[c][1]}" for c in sorted_classes])
    score_segments = ", ".join([f"'Class {c}'!E$6:E${class_sheet_ranges[c][1]}" for c in sorted_classes])
    pass_rate_segments = ", ".join([f"'Class {c}'!H${class_sheet_ranges[c][1]+1}" for c in sorted_classes])

    ws_dash["B6"] = f"=COUNTA({total_stud_segments})"
    ws_dash["B7"] = f"=AVERAGE({score_segments})"
    ws_dash["B7"].number_format = '0.0'
    ws_dash["B8"] = f"=MEDIAN({score_segments})"
    ws_dash["B8"].number_format = '0.0'
    ws_dash["B9"] = f"=MAX({score_segments})"
    ws_dash["B10"] = f"=MIN({score_segments})"
    ws_dash["B11"] = f"=AVERAGE({pass_rate_segments})"
    ws_dash["B11"].number_format = '0.0%'

    # Populate Class Breakdown rows dynamically
    for idx, c in enumerate(sorted_classes, start=15):
        s_row, e_row = class_sheet_ranges[c]
        ws_dash.cell(row=idx, column=1, value=c)
        ws_dash.cell(row=idx, column=2, value=f"=COUNTA('Class {c}'!B$6:B${e_row})")
        ws_dash.cell(row=idx, column=3, value=f"='Class {c}'!E${e_row+1}")
        ws_dash.cell(row=idx, column=4, value=f"=MAX('Class {c}'!E$6:E${e_row})")
        ws_dash.cell(row=idx, column=5, value=f"='Class {c}'!H${e_row+1}")
        
        for col_i in range(1, 6):
            cell = ws_dash.cell(row=idx, column=col_i)
            cell.font = regular_font; cell.border = thin_border
            if col_i > 1:
                cell.alignment = Alignment(horizontal="right")
                if col_i == 3: cell.number_format = '0.0'
                elif col_i == 5: cell.number_format = '0.0%'

    # Generate Dashboard Chart
    chart = BarChart()
    chart.type = "col"; chart.title = "Average Class Performance"
    chart.y_axis.title = "Score / 10"; chart.x_axis.title = "Class"
    chart.width = 15; chart.height = 10; chart.legend = None
    
    data_ref = Reference(ws_dash, min_col=3, min_row=14, max_row=14 + len(sorted_classes))
    cats_ref = Reference(ws_dash, min_col=1, min_row=15, max_row=14 + len(sorted_classes))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_dash.add_chart(chart, "G3")

    # ----------------------------------------------------
    # 6. TAB 3: ITEM ANALYSIS (COHORT DIAGNOSTIC)
    # ----------------------------------------------------
    ws_item = wb.create_sheet(title="Item Analysis")
    ws_item.views.sheetView[0].showGridLines = True
    ws_item["A1"] = "Cohort Item Analysis & Question Diagnostic"
    ws_item["A1"].font = title_font
    
    item_headers = ["Question", "Expected Answer", "Correct Count", "Incorrect Count", "Blank Count", "Ambiguous Count", "Success Rate"]
    for c_idx, h in enumerate(item_headers, start=1):
        cell = ws_item.cell(row=3, column=c_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

    for q_idx, q in enumerate(active_qs, start=4):
        q_letter = get_column_letter(8 + q) # Maps matching question column letter
        
        ws_item.cell(row=q_idx, column=1, value=f"Q{q}").font = bold_font
        ws_item.cell(row=q_idx, column=2, value=expected_answers.get(q, '')).alignment = Alignment(horizontal="center")
        
        # Build cross-tab cross-counting formulas dynamically
        correct_f = " + ".join([f"COUNTIF('Class {c}'!{q_letter}$6:{q_letter}${class_sheet_ranges[c][1]}, \"{expected_answers.get(q)}\")" for c in sorted_classes])
        blank_f = " + ".join([f"COUNTIF('Class {c}'!{q_letter}$6:{q_letter}${class_sheet_ranges[c][1]}, \"BLANK\")" for c in sorted_classes])
        ambig_f = " + ".join([f"COUNTIF('Class {c}'!{q_letter}$6:{q_letter}${class_sheet_ranges[c][1]}, \"AMBIG\")" for c in sorted_classes])
        total_stud_f = " + ".join([f"COUNTA('Class {c}'!B$6:B${class_sheet_ranges[c][1]})" for c in sorted_classes])
        
        ws_item.cell(row=q_idx, column=3, value=f"={correct_f}").alignment = Alignment(horizontal="right")
        ws_item.cell(row=q_idx, column=5, value=f"={blank_f}").alignment = Alignment(horizontal="right")
        ws_item.cell(row=q_idx, column=6, value=f"={ambig_f}").alignment = Alignment(horizontal="right")
        ws_item.cell(row=q_idx, column=4, value=f"=({total_stud_f}) - C{q_idx} - E{q_idx} - F{q_idx}").alignment = Alignment(horizontal="right")
        
        succ_cell = ws_item.cell(row=q_idx, column=7, value=f"=C{q_idx}/({total_stud_f})")
        succ_cell.alignment = Alignment(horizontal="right"); succ_cell.number_format = '0.0%'
        
        for c_idx in range(1, len(item_headers) + 1):
            cell = ws_item.cell(row=q_idx, column=c_idx); cell.border = thin_border
            if q_idx % 2 == 1: cell.fill = zebra_fill

    # Item Analysis Horizontal Chart
    chart_item = BarChart()
    chart_item.type = "bar"; chart_item.title = "Question Success Rate Diagnostic"
    chart_item.x_axis.title = "Success Rate"; chart_item.y_axis.title = "Question"
    chart_item.width = 15; chart_item.height = 11; chart_item.legend = None
    
    data_ref_item = Reference(ws_item, min_col=7, min_row=3, max_row=3 + len(active_qs))
    cats_ref_item = Reference(ws_item, min_col=1, min_row=4, max_row=3 + len(active_qs))
    chart_item.add_data(data_ref_item, titles_from_data=True)
    chart_item.set_categories(cats_ref_item)
    ws_item.add_chart(chart_item, "I3")

    # ----------------------------------------------------
    # 7. FORMAT PIPELINE (AUTOFIT & FREEZE PANES)
    # ----------------------------------------------------
    for sheet in wb.worksheets:
        if "Class" in sheet.title:
            sheet.freeze_panes = "I6" # Freeze right before individual answers
        elif "Item" in sheet.title:
            sheet.freeze_panes = "A4"
            
        for col in sheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    
                    # Fallback to 1 if .column is None, satisfying the type checker
                    col_idx = col[0].column if col[0].column is not None else 1
                    col_letter = get_column_letter(col_idx)
                    
                    sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(output_filename)
    print(f"Interactive Excel Dashboard successfully compiled and saved to '{output_filename}'")


generate_excel_dashboard(results_sorted)
